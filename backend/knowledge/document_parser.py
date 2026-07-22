"""文档解析器 - 支持txt/md/pdf/docx/xlsx"""
import os

def parse_file(filepath: str) -> str:
    """解析文件为纯文本"""
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext in (".txt", ".md", ".csv"):
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    
    elif ext == ".pdf":
        try:
            from pypdf2 import PdfReader
            reader = PdfReader(filepath)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            return text
        except ImportError:
            return "[PDF解析需要pypdf2，请pip install pypdf2]"
    
    elif ext == ".docx":
        try:
            from docx import Document
            doc = Document(filepath)
            return "\n".join([p.text for p in doc.paragraphs])
        except ImportError:
            return "[DOCX解析需要python-docx，请pip install python-docx]"
    
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n## {sheet}\n"
                for row in ws.iter_rows(values_only=True):
                    text += " | ".join([str(c or "") for c in row]) + "\n"
            return text
        except ImportError:
            return "[XLSX解析需要openpyxl]"
    
    return f"[不支持的文件格式: {ext}]"

def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list:
    """将文本分块用于embedding"""
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk.strip())
        start = end - overlap
    return chunks
